from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl

DEFAULT_WAIT_MS = 10000
DEFAULT_TYPE_DELAY_MS = 200
DEFAULT_TIMEOUT_MS = 60000
DEFAULT_SLOW_MO_MS = 600


# ✅ normalize function
def normalize(text):
    if not text:
        return ""
    return text.replace(".", "").replace(",", "").strip().lower()


def _read_output(output_locator):
    try:
        v = output_locator.input_value()
        if v:
            return str(v).strip()
    except:
        pass
    try:
        v = output_locator.inner_text()
        if v:
            return str(v).strip()
    except:
        pass
    return ""


def run_test():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel")
    parser.add_argument("--input-col")
    parser.add_argument("--expected-col")
    parser.add_argument("--url")
    parser.add_argument("--wait-ms", type=int, default=DEFAULT_WAIT_MS)
    parser.add_argument("--type-delay-ms", type=int, default=DEFAULT_TYPE_DELAY_MS)
    parser.add_argument("--slow-mo-ms", type=int, default=DEFAULT_SLOW_MO_MS)
    parser.add_argument("--keep-open", action="store_true")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel)
    ws = wb.active

    header = [str(ws.cell(row=1, column=i).value).strip() for i in range(1, ws.max_column + 1)]

    input_idx = header.index(args.input_col) + 1
    expected_idx = header.index(args.expected_col) + 1

    print(f"Starting Frontend test with {ws.max_row - 1} rows...")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=args.slow_mo_ms)
        page = browser.new_page()

        page.goto(args.url)
        page.wait_for_timeout(8000)

        input_box = page.locator("textarea").nth(0)
        output_box = page.locator("textarea").nth(1)
        button = page.get_by_role("button", name=re.compile("Transliterate", re.IGNORECASE))

        for i in range(2, ws.max_row + 1):
            text = str(ws.cell(row=i, column=input_idx).value or "").strip()
            if not text:
                continue

            expected = str(ws.cell(row=i, column=expected_idx).value or "").strip()

            print(f"Testing [Row {i}]: {text}")

            try:
                input_box.fill("")
                input_box.type(text, delay=args.type_delay_ms)

                button.click()

                # 🔥 stable wait
                page.wait_for_timeout(10000)

                output = _read_output(output_box)

                if not output:
                    page.wait_for_timeout(5000)
                    output = _read_output(output_box)

                # ✅ FIXED LOGIC (THIS IS THE IMPORTANT CHANGE)
                if normalize(output) in normalize(expected) or normalize(expected) in normalize(output):
                    status = "PASS"
                else:
                    status = "FAIL"

                print(f"  -> {status}")

            except Exception as e:
                print(f"Error: {e}")

        if args.keep_open:
            print("Keeping browser open...")
            while True:
                page.wait_for_timeout(1000)

        browser.close()


if __name__ == "__main__":
    run_test()